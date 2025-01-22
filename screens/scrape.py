import openpyxl
import streamlit as st # type: ignore
from screens import logout
from openpyxl.styles import PatternFill
from data import extrac_from_db, dowload_results_file

from scrappers import (
    check_amazon, 
    check_liverpool,
    check_mercadolibre, 
    check_walmart,
    check_home_depot,
    check_coppel
)

def scrape_page(supabase):
    if st.button("Cerrar Sesión"):
        logout(supabase=supabase)
    
    st.title("Marketplace Product Status Extractor")

    st.header("Extrac data from database", divider="blue")

    column1 = st.columns(1)
    col1, col2, col3 = st.columns(3)
    col21, col22, col23 = st.columns(3)
    if column1[0].button("Descargar todas las marketplaces"):
        extrac_from_db(marketplace="All", supabase=supabase,isSendingByEmail=False)

    if col1.button("Amazon"):
        extrac_from_db(marketplace="Amazon", supabase=supabase, isSendingByEmail=False)

    if col2.button("Mercado Libre"):
        extrac_from_db(marketplace="MercadoLibre", supabase=supabase, isSendingByEmail=False)

    if col3.button("Liverpool"):
        extrac_from_db(marketplace="Liverpool", supabase=supabase, isSendingByEmail=False)

    if col21.button("Walmart"):
        extrac_from_db(marketplace="Walmart", supabase=supabase, isSendingByEmail=False)

    if col22.button("Home Depot"):
        extrac_from_db(marketplace="HomeDepot", supabase=supabase, isSendingByEmail=False)

    if col23.button("Coppel"):
        extrac_from_db(marketplace="Coppel", supabase=supabase, isSendingByEmail=False)


    st.header("Extrac data from excel file", divider="green")

    dataset = st.file_uploader("Upload Excel file (.xlsx)", type=['xlsx'])

    if dataset is not None:
        build_excel_results(dataset=dataset)
        

def build_excel_results(dataset):

    wb = openpyxl.load_workbook(dataset, read_only=True)
    st.info(f"File uploaded: {dataset.name}")

    ws = wb.active

    ###
    # End Result Excel Variables
    result_wb = openpyxl.Workbook()
    result_ws = result_wb.active

    keys = [
        "Marketplace", "Codigo", "Descripcion", "Link", "Estatus",
        "Precio Regular", "Precio Promoción", "Calificación", "# Reseñas"
    ]
    result_row_num = 1

    for col_num, column_title in enumerate(keys, 1):
        cell = result_ws.cell(row=result_row_num, column=col_num) # type: ignore
        cell.value = column_title

    st.write("Procesando archivo...")
    prg = st.progress(0)

    i = 0
    for row in ws.iter_rows(min_row=2, values_only=True):# type: ignore
        total = ws.max_row - 1
        if row[0] is None:
            continue
        result_row_num += 1

        marketplace = row[0]
        product_code = row[1]
        product_name = row[2]
        link = row[3]

        result = ""
        price = "-"
        rating = "-"
        reviews = "-"
        promotion_price = "-"
        if marketplace == 'Amazon':
            result, price, promotion_price, rating, reviews = check_amazon(
                link)
        elif marketplace == 'ML':
            result, price, promotion_price, rating, reviews = check_mercadolibre(
                link)

        elif marketplace == 'Liverpool':
            result, price, promotion_price, rating, reviews = check_liverpool(
                link)
        elif marketplace == 'Walmart':
            result, price, promotion_price, rating, reviews = check_walmart(link)# type: ignore
        elif marketplace == 'HomeDepot':
            result, price, promotion_price, rating, reviews = check_home_depot(
                link)
        elif marketplace == 'Coppel':
            result, price, promotion_price, rating, reviews = check_coppel(
                link)

        row = [
            marketplace, product_code, product_name, link, result, price,
            promotion_price, rating, reviews
        ]

        for col_num, cell_value in enumerate(row, 1):
            cell = result_ws.cell(row=result_row_num, column=col_num)# type: ignore
            cell.value = cell_value
        i += 1
        prg.progress(i / total,text=f"Procesando {i} producto(s) de {total}.")

    st.write("Terminando de procesar archivo...")
    st.write("Se procesaron ", i, " productos.")
    for e_column in result_ws['E']:# type: ignore
        if e_column.value == "ACTIVO":
            e_column.fill = PatternFill(start_color='38B856',
                                        end_color='38B856',
                                        fill_type='solid')
        if e_column.value == "INACTIVO":
            e_column.fill = PatternFill(start_color='d30000',
                                        end_color='d30000',
                                        fill_type='solid')
        if e_column.value in [
                "PAGINA NO ENCONTRADA", "Failed to fetch the page"
        ]:
            e_column.fill = PatternFill(start_color='808080',
                                        end_color='808080',
                                        fill_type='solid')
    
    dowload_results_file(marketplace=marketplace,result_wb=result_wb)