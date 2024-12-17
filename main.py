import requests
from bs4 import BeautifulSoup
import streamlit as st # type: ignore
import pandas as pd
import openpyxl
from io import BytesIO
from tempfile import NamedTemporaryFile
from openpyxl.styles import Color, PatternFill
import json
import decimal
from itertools import cycle
import random
import time
import os
from dotenv import load_dotenv # type: ignore
from firecrawl import FirecrawlApp  # type: ignore
import re

from scrappers import check_amazon, check_liverpool,check_mercadolibre



def chunk_list(lst, chunk_size):
    """Splits a list into chunks of a specific size."""
    return [lst[i:i + chunk_size] for i in range(0, len(lst), chunk_size)]

def check_walmart(url):
    try:
        load_dotenv()

        app = FirecrawlApp(api_key=os.getenv('FC_API_KEY'))

        time.sleep(random.uniform(1, 4))

        response = app.scrape_url(url=url, params={
            'formats': [ 'markdown' ],
        })

        current_price = None
        original_price = None
        stars = None
        reviews = None

        current_price_match = re.search(r'MXN$(\d+\.\d{2})', response["markdown"])
        current_price = f"${current_price_match.group(1)}" if current_price_match else None

        original_price_match = re.search(r'costaba \$(\d+\.\d{2})', response["markdown"])
        original_price = f"${original_price_match.group(1)}" if original_price_match else None

        stars_pattern = r"\((\d+\.\d+)\)(\d+\.\d+) estrellas de (\d+) reseñas"
        match_stars = re.search(stars_pattern, response["markdown"])

        if match_stars:
            stars = match_stars.group(1)
            reviews = match_stars.group(3)
        else:
            stars = None
            reviews = None

        if current_price is not None:
            return ("ACTIVO",
                    current_price,
                    (original_price
                        if original_price is not None else "-"),
                    (stars
                        if stars is not None else "-"),
                    (reviews
                        if reviews is not None else "-"))
        else:
            price_pattern = r"MXN\$(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)"
            match = re.search(price_pattern, response["markdown"])
            if match:
                current_price = f"${match.group(1)}"
                return ("ACTIVO",
                    current_price,
                    "-",
                     (stars
                        if stars is not None else "-"),
                    (reviews
                        if reviews is not None else "-"))
            else:
                return "INACTIVO", "-", "-", "-", "-"

    except requests.RequestException as e:
        return "PAGINA NO ENCONTRADA", 0, 0, "-", "-"



def check_home_depot(url):
    try:

        last_part = url.split("/")[-1]
        last_value = last_part.split("-")[-1]
        home_depot_request = f'https://www.homedepot.com.mx/search/resources/api/v2/products?langId=-5&storeId=10351&contractId=4000000000000000003&langId=-5&partNumber={last_value}&physicalStoreId=8702'
        response = requests.get(home_depot_request)
        data = response.json()
        promotion_price = None
        rating = None
        review = None
        list_price = None
        status = "INACTIVO"
        base_data = data["contents"][0]
        for price_pos in base_data["price"]:
            if price_pos["usage"] == "Offer":
                promotion_price = price_pos["value"]
            if price_pos["usage"] == "Display":
                list_price = price_pos["value"]
        if "x_ratings.total_reviews" in base_data:
            review = base_data["x_ratings.total_reviews"]
        if "x_ratings.rating" in base_data:
            rating = base_data["x_ratings.rating"]
        product_id = base_data["id"]
        status_request = f'https://www.homedepot.com.mx/wcs/resources/store/10351/inventoryavailability/{product_id}?&langId=-5&onlineStoreId=10351&search=2&physicalStoreId=12605'
        response = requests.get(status_request)
        availabilty_data = response.json()
        if availabilty_data["InventoryAvailability"][0][
                "inventoryStatus"] == "Available":
            status = "ACTIVO"

        return (status, (list_price if list_price is not None else "-"),
                (promotion_price if promotion_price is not None else "-"),
                (rating if rating is not None else "-"),
                (review if review is not None else "-"))
    except requests.RequestException:
        return "Failed to fetch the page"

def check_coppel(url):
    #url must be https://www.coppel.com.mx/$product
    try:
        response = requests.get(url)
        if response.status_code == 200:
            discounted_price = None
            original_price = None
            soup = BeautifulSoup(response.text, 'html.parser')

            discounted_price = soup.find('h2', {'data-testid': 'pdp_discounted_price'})

            if discounted_price is not None:
                original_price = soup.find('p', {'data-testid': 'pdp_price'})
            else:
                original_price = soup.find('h2', {'data-testid': 'pdp_price'})


            if discounted_price is None and original_price is None:
                return "PAGINA NO ENCONTRADA", "-", "-", "-", "-"

            return "ACTIVO", (original_price.text if original_price is not None else "-"), (discounted_price.text if discounted_price is not None else "-"),  "-", "-"
        else:
            return "INACTIVO", "-", "-", "-", "-"
    except requests.RequestException as e:
            return "PAGINA NO ENCONTRADA", "-", "-", "-", "-"

def main():
    # mas de un archivo
    # descarga del zip creado de facturapi
    st.title("Marketplace Product Status Extractor")

    dataset = st.file_uploader("Upload Excel file (.xlsx)", type=['xlsx'])
    results = {}

    if dataset is not None:
        wb = openpyxl.load_workbook(dataset, read_only=True)
        st.info(f"File uploaded: {dataset.name}")

        ws = wb.active

        ###
        # End Result Excel Variables
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active

        keys = [
            "Marketplace", "Codigo", "Descripcion", "Link", "Estatus",
            "Precio Regular", "Precio Promoción", "Calificación", "# Reseñas"
        ]
        result_row_num = 1

        for col_num, column_title in enumerate(keys, 1):
            cell = result_ws.cell(row=result_row_num, column=col_num) # type: ignore
            cell.value = column_title

        st.write("Procesando archivo...")

        i = 0
        for row in ws.iter_rows(min_row=2, values_only=True):# type: ignore
            if row[0] is None:
                continue
            result_row_num += 1

            marketplace = row[0]
            product_code = row[1]
            product_name = row[2]
            link = row[3]

            result = ""
            price = "-"
            rating = "-"
            reviews = "-"
            promotion_price = "-"
            if marketplace == 'Amazon':
                result, price, promotion_price, rating, reviews = check_amazon(
                    link)
            elif marketplace == 'ML':
                result, price, promotion_price, rating, reviews = check_mercadolibre(
                    link)

            elif marketplace == 'Liverpool':
                result, price, promotion_price, rating, reviews = check_liverpool(
                    link)
            elif marketplace == 'Walmart':
                result, price, promotion_price, rating, reviews = check_walmart(link)# type: ignore
            elif marketplace == 'HomeDepot':
                result, price, promotion_price, rating, reviews = check_home_depot(
                    link)
            elif marketplace == 'Coppel':
                result, price, promotion_price, rating, reviews = check_coppel(
                    link)

            row = [
                marketplace, product_code, product_name, link, result, price,
                promotion_price, rating, reviews
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = result_ws.cell(row=result_row_num, column=col_num)# type: ignore
                cell.value = cell_value
            i += 1

        st.write("Terminando de procesar archivo...")
        st.write("Se procesaron ", i, " productos.")
        for e_column in result_ws['E']:# type: ignore
            if e_column.value == "ACTIVO":
                e_column.fill = PatternFill(start_color='38B856',
                                            end_color='38B856',
                                            fill_type='solid')
            if e_column.value == "INACTIVO":
                e_column.fill = PatternFill(start_color='d30000',
                                            end_color='d30000',
                                            fill_type='solid')
            if e_column.value in [
                    "PAGINA NO ENCONTRADA", "Failed to fetch the page"
            ]:
                e_column.fill = PatternFill(start_color='808080',
                                            end_color='808080',
                                            fill_type='solid')
        with NamedTemporaryFile() as tmp:
            result_wb.save(tmp.name)
            data = BytesIO(tmp.read())

        st.subheader("Resultados")
        st.download_button("Descargar Archivo",
                           data=data,
                           mime='xlsx',
                           file_name="resultados.xlsx")


if __name__ == "__main__":
    main()
