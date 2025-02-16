import openpyxl
from io import BytesIO
import streamlit as st # type: ignore
from tempfile import NamedTemporaryFile
from openpyxl.styles import PatternFill
from scrappers import (
    check_amazon, 
    check_liverpool,
    check_mercadolibre, 
    check_walmart,
    check_home_depot,
    check_coppel
)
from base64 import b64encode
import os

def dowload_results_file(file_name,result_wb,isSendingByEmail):
    try:
        with NamedTemporaryFile(delete=False) as tmp:
            result_wb.save(tmp.name)

            if isSendingByEmail:
                tmp.seek(0)
                file_data = tmp.read()
                encoded_file = b64encode(file_data).decode()
                print("encoded_file[:100]")
                print(encoded_file[:100])
                return encoded_file
            else:
                data = BytesIO(tmp.read())
                st.subheader("Resultados")
                st.download_button("Descargar Archivo",
                                    data=data,
                                    mime='xlsx',
                                    file_name=f"resultados_{file_name}.xlsx")
    
    except Exception as e:
        print(f"Error al generar o leer el archivo: {str(e)}")
        return None
    finally:
        if isSendingByEmail or os.path.exists(tmp.name):
            os.remove(tmp.name)

def extrac_from_db(marketplace,supabase,isSendingByEmail):
    try:
        if not isSendingByEmail:
            st.info(f"Processing data from {marketplace}...")

        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active
        file_name = marketplace

        keys = [
            "Marketplace", "Codigo", "Descripcion", "Link", "Estatus",
            "Precio Regular", "Precio Promoción", "Calificación", "# Reseñas"
        ]
        result_row_num = 1

        for col_num, column_title in enumerate(keys, 1):
            cell = result_ws.cell(row=result_row_num, column=col_num) # type: ignore
            cell.value = column_title

        i = 0
        all_products = []
        if marketplace == 'All':
            limit = 1000 # default limit for supabase requests
            offset = 0
            file_name = "todas_marketplaces"

            while True:
                resp = supabase.table("Products").select("*").eq("active", 1).range(offset, offset + limit - 1).execute()
                all_products.extend(resp.data)
                if len(resp.data) < limit:
                    break

                offset += limit
        else:
            resp = supabase.table("Products").select("*").eq("marketplace",marketplace).eq("active",1).execute()
            all_products.extend(resp.data)
        if not isSendingByEmail:
            prg = st.progress(0)
        
        if resp.data:
            total = len(all_products)
            for product in all_products:
                result_row_num += 1
                product_code = product["code"]
                product_name = product["name"]
                marketplace = product["marketplace"]
                link = product["link"]
                result = ""
                price = "-"
                rating = "-"
                reviews = "-"
                promotion_price = "-"

                if marketplace == 'Amazon':
                    result, price, promotion_price, rating, reviews = check_amazon(
                        str(link))
                elif marketplace == 'MercadoLibre':
                    result, price, promotion_price, rating, reviews = check_mercadolibre(
                        str(link))

                elif marketplace == 'Liverpool':
                    result, price, promotion_price, rating, reviews = check_liverpool(
                        str(link))
                elif marketplace == 'Walmart':
                    result, price, promotion_price, rating, reviews = check_walmart(str(link))# type: ignore
                elif marketplace == 'HomeDepot':
                    result, price, promotion_price, rating, reviews = check_home_depot(
                        str(link))
                elif marketplace == 'Coppel':
                    result, price, promotion_price, rating, reviews = check_coppel(
                        str(link))
                    
                row = [
                    marketplace, product_code, product_name, link, result, price,
                    promotion_price, rating, reviews
                ]

                for col_num, cell_value in enumerate(row, 1):
                    cell = result_ws.cell(row=result_row_num, column=col_num)# type: ignore
                    cell.value = cell_value
                i += 1
                if not isSendingByEmail:
                    prg.progress(i / total,text=f"Procesando {i} producto(s) de {total}.")
                else:
                    print(f"Procesando {i} producto(s) de {total}.")
            if not isSendingByEmail:
                st.write("Terminando de procesar los datos...")
                st.write("Se procesaron ", i, " productos.")
            for e_column in result_ws['E']:# type: ignore
                if e_column.value == "ACTIVO":
                    e_column.fill = PatternFill(start_color='38B856',
                                                end_color='38B856',
                                                fill_type='solid')
                if e_column.value == "INACTIVO":
                    e_column.fill = PatternFill(start_color='d30000',
                                                end_color='d30000',
                                                fill_type='solid')
                if e_column.value in [
                        "PAGINA NO ENCONTRADA", "Failed to fetch the page"
                ]:
                    e_column.fill = PatternFill(start_color='808080',
                                                end_color='808080',
                                                fill_type='solid')
            if isSendingByEmail:
                return dowload_results_file(file_name=file_name,result_wb=result_wb,isSendingByEmail=isSendingByEmail)
            else:        
                dowload_results_file(file_name=file_name,result_wb=result_wb,isSendingByEmail=isSendingByEmail)

        else:
            if not isSendingByEmail:
                st.warning(f"No data in table {marketplace}.")
    except Exception as e:
        if not isSendingByEmail:
            st.error(f"Error to obtain products of {marketplace}: {e}")
            if st.button("Reiniciar proceso",type="secondary"):
                st.rerun()

        return []
    